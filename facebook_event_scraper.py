# -*- coding: utf-8 -*-
import os
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import requests
import datetime
from dateutil import parser
import json
from ignoreconstants import ignore_openpyxl_constants                                               
ignore_openpyxl_constants()
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Style, Font

#This is for building .exe with cx_freeze otherwise cx_freeze can't find the cacert.pem
from sys import platform as _platform
if _platform == "win32":
    os.environ["REQUESTS_CA_BUNDLE"] = os.path.join(os.getcwd(), "cacert.pem")

#id,name,description,ticket_uri,start_time,place,end_time
__author__ = "Ferhat Özmen"
__copyright__ = "Copyright 2016"
__license__ = "GPL"
__version__ = "0.1"
__maintainer__ = "Ferhat Özmen"
__email__ = "sreungbrmzra@gmail.com"

token = '' #Set your token here

currentTime = datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
EXCELFILE = 'event_list.xlsx'

#Excel Row Names
ROW_EVENT_NAME = 'Event Name'
ROW_FACEBOOK_EVENT_URL = 'Facebook Event URL'
ROW_DATE_START = 'Date of Event'
ROW_LOCATION = 'Location'
ROW_TICKET_URI = 'Ticket Link'
ROW_INTERESTED_COUNT = 'Interested Count'
ROW_GOING_COUNT = 'Going Count'
ROW_CITY = 'City'

def saveEvent(event, wb, ws):
    event = getEvent(event['id']).json()
    event_name = event['name']
    event_facebook_url = "https://www.facebook.com/events/%s" %event['id']
    date_start = parser.parse(event['start_time'])
    ticket_uri = event['ticket_uri']
    interested_count = event['interested_count']
    attending_count = event['attending_count']
    city = event['place']['location']['city']

    print event_name.encode('utf-8')
    print event_facebook_url.encode('utf-8')
    print date_start
    print city.encode('utf-8')
    print ticket_uri.encode('utf-8')
    print interested_count
    print attending_count
    print "************************************************"

    row_number = ws.max_row+1
    ws.cell(row=row_number, column=1, value=event_name)
    ws.cell(row=row_number, column=2, value=event_facebook_url)
    ws.cell(row=row_number, column=3, value=date_start)
    ws.cell(row=row_number, column=4, value=ticket_uri)
    ws.cell(row=row_number, column=5, value=city)
    ws.cell(row=row_number, column=6, value=interested_count)
    ws.cell(row=row_number, column=7, value=attending_count)
    wb.save(EXCELFILE)

def createExcelFile():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=ROW_EVENT_NAME).font = Font(bold=True)
    ws.cell(row=1, column=2, value=ROW_FACEBOOK_EVENT_URL).font = Font(bold=True)
    ws.cell(row=1, column=3, value=ROW_DATE_START).font = Font(bold=True)
    ws.cell(row=1, column=4, value=ROW_TICKET_URI).font = Font(bold=True)
    ws.cell(row=1, column=5, value=ROW_CITY).font = Font(bold=True)
    ws.cell(row=1, column=6, value=ROW_INTERESTED_COUNT).font = Font(bold=True)
    ws.cell(row=1, column=7, value=ROW_GOING_COUNT).font = Font(bold=True)
    wb.save(EXCELFILE)

    return wb, ws

def getSession():
    session = requests.Session()
    return session

def getGraphAPICall(token, query, time):
    url = "https://graph.facebook.com/v2.7/search?"
    payload = {'access_token': token, 'q': query, 'type': 'event', 'limit': '1000', 'since': time,
               'fields': 'city,state,country,description,id,start_time,end_time,name,place,street,\
               zip,ticket_uri,intested_count,attending_count'}
    return url, payload

def getEvent(event_id):
    url = "https://graph.facebook.com/v2.7/%s?" %event_id
    payload = {'access_token': token, 'fields': 'name,id,attending_count,\
               interested_count,ticket_uri,place,start_time,end_time'}
    result = requests.get(url, params=payload)
    return result

def makeGraphAPICall(url, payload, session):
    results = session.get(url, params=payload) 
    return results

def getNextPage(next_page, session):
    results = session.get(next_page)
    return results

def validateEvent(event, query, id_list):
    if ('place' in event and 'location' in event['place'] and\
            'city' in event['place']['location'] and 'ticket_uri' in event and\
            event['place']['location']['city'] == query and\
                    event['id'] not in id_list):
        id_list.append(event['id'])
        return True
    else:
        return False

def parseResults(results, session, query, id_list, wb, ws):
    results = results.json()
    while 'paging' in results:
        for event in results['data']:
            #saveEvent(event)
            if validateEvent(event, query, id_list):
                saveEvent(event, wb, ws)
        results = getNextPage(results['paging']['next'], session).json() 

def getExcelSheet():
    if os.path.isfile(EXCELFILE):
        print "The Excel Sheet: %s does already exist, delete or put it somewhere else please!"\
                %EXCELFILE
        sys.exit()
    else:
        wb, ws = createExcelFile()
        return wb, ws

def main():
    id_list = []
    query = raw_input("Please enter the City: ").title()
    session = getSession()
    wb, ws = getExcelSheet()
    url, payload = getGraphAPICall(token, query, currentTime)
    results = makeGraphAPICall(url, payload, session)
    parseResults(results, session, query, id_list, wb, ws)

    print "################################"
    print "Finished Scraping Events for the query: %s" %query
    print "Total scraped Events: %i" %len(id_list)

def test(event_id):
    print json.dumps(getEvent(event_id).json(), indent=1)

if __name__ == '__main__':
    main()
    #test('171253766579392')
